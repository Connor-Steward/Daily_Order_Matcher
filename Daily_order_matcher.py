import openpyxl

def highlight_matching_columns(input_file1, input_file2, output_file):
    wb1 = openpyxl.load_workbook(input_file1)
    sheet1 = wb1["Sheet1"]

    wb2 = openpyxl.load_workbook(input_file2)
    sheet2 = wb2["Sheet1"]

    print("Started comparing the files...")

    for i, row1 in enumerate(sheet1.iter_rows(values_only=True)):
        if i == 0:  # Skip the first row
            continue

        #print(row1)

        for j, row2 in enumerate(sheet2.iter_rows(values_only=True)):
            if j == 0:  # Skip the first row
                continue
            if row1[3] == row2[3] and row1[8] == row2[8]:
                # Row1/2[3] = Serial Column
                # Row1/2[8] = Desc Column
                print("Matched row")
                print(row1)
                print(j)
                # Column 20 + row1[20] is Column T in Excel
                sheet2.cell(row=j+1, column=20).value = row1[19]


    print("Finished comparing the files...")
    wb2.save(output_file)
    print("Saved the output file")





highlight_matching_columns("sheet1.xlsx", "sheet2.xlsx", "output_file.xlsx")
